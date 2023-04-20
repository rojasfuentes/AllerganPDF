from same import matches_noCliente_list, nombreCliente_list, direccionCliente_list, colCliente_list, zipcodeCliente_list, ciudad_cliente_list, consignatario_list, direccionConsig_list, colConsig_list, ciudad_consig_list, zipcodeConsig_list, clave_list, descripcion_list, cantidad_list, lote_list

import openpyxl

# CRAR UN ARCHIVO NUEVO
wb = openpyxl.Workbook()

# Selecciona la hoja de trabajo
ws = wb.active

# Combinar celdas para "Bill To" en B1 y C1
ws.merge_cells('B1:C1')
ws['B1'] = 'Bill To'

ws.merge_cells('F1:G1')
ws['F1'] = 'Ship To'

# Ajustar el ancho de la columna
ws.column_dimensions['B'].width = 15

# Agregar "Cliente" en B2 y F2
ws['B2'] = 'Cliente'
ws['C2'] = matches_noCliente

ws['B3'] = 'Nombre'
ws['C3'] = nombreCliente

ws['B4'] = 'Dirección'
ws['C4'] = direccionCliente

ws['B5'] = 'Colonia'
ws['C5'] = colCliente

ws['B6'] = 'C.P.'
ws['C6'] = zipcodeCliente

ws['B7'] = 'Ciudad'
ws['C7'] = ciudad_cliente


# Consignatario
ws['B3'] = 'Nombre'
ws['G3'] = consignatario

ws['F4'] = 'Dirección'
ws['G4'] = direccionConsig

ws['F5'] = 'Colonia'
ws['G5'] = colConsig

ws['F6'] = 'C.P.'
ws['G6'] = zipcodeConsig

ws['F7'] = 'Ciudad'
ws['G7'] = ciudad_consig


# Datos producto
ws['B11'] = 'Clave'
ws['C11'] = clave

ws['B12'] = 'Lote'
ws['C12'] = lote

ws['B13'] = 'Cantidad'
ws['C13'] = cantidad

ws['B14'] = 'Descripción'
ws['C14'] = descripcion

# Ajustar el ancho de las columnas de forma automática
for column_cells in ws.columns:
    column = column_cells[0].column
    adjusted_width = max(len(str(cell.value)) for cell in column_cells) + 2
    column_letter = openpyxl.utils.get_column_letter(column)
    # establece el ancho ajustado de la columna
    ws.column_dimensions[column_letter].width = adjusted_width


# Guardar el archivo de excel actualizado
wb.save('C:/Users/JFROJAS/Desktop/CarlosPDF/project/resultados.xlsx')
