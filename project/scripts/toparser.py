from same import matches_noCliente_list, nombreCliente_list, direccionCliente_list, colCliente_list, zipcodeCliente_list, ciudad_cliente_list, consignatario_list, direccionConsig_list, colConsig_list, ciudad_consig_list, zipcodeConsig_list, clave_list, descripcion_list, cantidad_list, lote_list
import openpyxl

# Cargar el libro de trabajo existente
workbook = openpyxl.load_workbook(filename=r'C:\Users\JFROJAS\Desktop\CarlosPDF\Carga Pedidos.xlsm', read_only=False, keep_vba=True)

# Crear una nueva hoja
new_sheet = workbook.create_sheet("Nueva Hoja")

# Escribir los datos en la nueva hoja
for i, pedido in enumerate(matches_noCliente_list):
    new_sheet.cell(row=i+1, column=1, value=pedido)
    new_sheet.cell(row=i+1, column=2, value=descripcion_list[i])
    new_sheet.cell(row=i+1, column=3, value=clave_list[i])
    new_sheet.cell(row=i+1, column=4, value=lote_list[i])
    new_sheet.cell(row=i+1, column=5, value=cantidad_list[i])
    new_sheet.cell(row=i+1, column=6, value=nombreCliente_list[i])
    new_sheet.cell(row=i+1, column=7, value=consignatario_list[i])


# Guardar el libro de trabajo
workbook.save(r'C:\Users\JFROJAS\Desktop\CarlosPDF\Carga Pedidos.xlsm')
